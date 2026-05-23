"""2024版房屋建筑工程预算定额 Excel 导入器.

Supports two Excel layouts:
1. **Standard layout** (一表式) — one row per quota sub-item:
   columns: 定额编号 | 定额名称 | 单位 | 人工费 | 材料费 | 机械费 | 合计
   + optional: 章节 | 工作内容 | 适用范围 | 地区 | 备注

2. **Detailed layout** (资源明细式) — one row per resource under a parent quota,
   with a "type" column indicating 人工/材料/机械.

Usage::

    from app.services.quota_import_2024 import import_quota_excel
    with open("房建定额2024.xlsx", "rb") as f:
        result = import_quota_excel(
            file_bytes=f.read(),
            db=session,
            standard_code="GBT50500-2024",
            profession="房建",
            region="全国",
            override_version="2024房建预算定额",
        )
    print(result.imported, result.skipped, result.errors)
"""

from __future__ import annotations

import io
import json
import logging
from dataclasses import dataclass, field

import openpyxl
from sqlalchemy.orm import Session

from app.models.quota_item import QuotaItem
from app.models.pricing_standard import PricingStandard

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Header → field mapping
# ---------------------------------------------------------------------------

_HEADER_MAP: dict[str, str] = {
    # Quota code
    "定额编号": "quota_code", "编号": "quota_code", "quota_code": "quota_code",
    "编码": "quota_code", "子目编号": "quota_code",
    # Name
    "定额名称": "name", "项目名称": "name", "名称": "name", "子目名称": "name",
    # Unit
    "单位": "unit", "计量单位": "unit",
    # Fee amounts (2024 style — direct ¥ amounts per unit)
    "人工费": "labor_fee", "人工费(元)": "labor_fee", "人工费（元）": "labor_fee",
    "材料费": "material_fee", "材料费(元)": "material_fee", "材料费（元）": "material_fee",
    "机械费": "machine_fee", "机械费(元)": "machine_fee", "机械费（元）": "machine_fee",
    "综合单价": "base_price", "合计": "base_price", "定额综合单价": "base_price",
    # Legacy qty fields (2013/2018 style)
    "人工工日": "labor_qty", "人工(工日)": "labor_qty",
    "材料消耗": "material_qty",
    "机械台班": "machine_qty", "机械(台班)": "machine_qty",
    # Metadata
    "章节": "chapter", "所属章节": "chapter", "分部": "chapter",
    "工作内容": "work_content",
    "适用范围": "applicable_scope",
    "地区": "region",
    "版本": "version",
    "备注": "remark",
}


def _normalize_header(raw: str) -> str | None:
    h = raw.strip().lower().replace(" ", "").replace("　", "")
    # Try exact match first (case-insensitive)
    for k, v in _HEADER_MAP.items():
        if k.lower().replace(" ", "") == h:
            return v
    return None


# ---------------------------------------------------------------------------
# Result dataclass
# ---------------------------------------------------------------------------

@dataclass
class QuotaImportResult:
    imported: int = 0
    skipped: int = 0
    updated: int = 0
    errors: list[str] = field(default_factory=list)
    items: list[QuotaItem] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Main importer
# ---------------------------------------------------------------------------

def import_quota_excel(
    file_bytes: bytes,
    db: Session,
    standard_code: str = "GBT50500-2024",
    profession: str = "房建",
    region: str = "全国",
    override_version: str = "",
    upsert: bool = True,
    sheet_name: str | None = None,
) -> QuotaImportResult:
    """Parse a 2024 housing-construction quota Excel file and upsert QuotaItems.

    Args:
        file_bytes: Raw xlsx bytes.
        db: SQLAlchemy session.
        standard_code: pricing_standards.code to link to (looked up by code).
        profession: Quota profession tag (e.g. "房建", "市政").
        region: Province/region for labor cost differentiation.
        override_version: Optional version tag override; if empty, auto-generated.
        upsert: If True, update existing rows matched by quota_code. If False, skip.
        sheet_name: Specific sheet to import; defaults to first sheet.
    """
    result = QuotaImportResult()

    # Resolve pricing_standard_id
    std_row = db.query(PricingStandard).filter(PricingStandard.code == standard_code).first()
    std_id: int | None = std_row.id if std_row else None
    if not std_id:
        logger.warning("PricingStandard '%s' not found; items will have null standard_id", standard_code)

    version_tag = override_version or f"{standard_code}-{profession}"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        result.errors.append(f"无法解析 Excel 文件: {exc}")
        return result

    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    if ws is None:
        result.errors.append("找不到工作表")
        return result

    # ── Detect header row ────────────────────────────────────────────────────
    col_map: dict[int, str] = {}
    header_row_idx: int | None = None

    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
        hits = 0
        temp_map: dict[int, str] = {}
        for col_idx, cell_val in enumerate(row):
            if cell_val is None:
                continue
            field_name = _normalize_header(str(cell_val))
            if field_name:
                temp_map[col_idx] = field_name
                hits += 1
        if hits >= 3:  # require at least 3 recognised columns
            col_map = temp_map
            header_row_idx = row_idx
            break

    if not col_map:
        result.errors.append(
            "未找到有效表头行（需包含定额编号/名称/人工费等列），请检查文件格式"
        )
        return result

    # ── Parse data rows ──────────────────────────────────────────────────────
    current_chapter = ""
    rows_processed = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if all(v is None for v in row):
            continue  # skip blank rows

        data: dict[str, object] = {}
        for col_idx, field_name in col_map.items():
            if col_idx < len(row):
                data[field_name] = row[col_idx]

        quota_code = _str(data.get("quota_code"))
        name = _str(data.get("name"))

        if not quota_code or not name:
            # May be a chapter heading row — capture for subsequent items
            if name and not quota_code:
                current_chapter = name
            continue

        rows_processed += 1
        chapter_val = _str(data.get("chapter")) or current_chapter
        labor_fee = _float(data.get("labor_fee"))
        material_fee = _float(data.get("material_fee"))
        machine_fee = _float(data.get("machine_fee"))
        base_price = _float(data.get("base_price"))
        # If base_price not given, derive from fee sum
        if not base_price:
            base_price = round(labor_fee + material_fee + machine_fee, 2)

        # Attempt upsert
        existing = db.query(QuotaItem).filter(QuotaItem.quota_code == quota_code).first()
        if existing:
            if not upsert:
                result.skipped += 1
                continue
            # Update 2024-specific fields
            existing.name = name
            existing.unit = _str(data.get("unit")) or existing.unit
            existing.labor_fee = labor_fee
            existing.material_fee = material_fee
            existing.machine_fee = machine_fee
            existing.base_price = base_price
            existing.labor_qty = _float(data.get("labor_qty")) or existing.labor_qty
            existing.material_qty = _float(data.get("material_qty")) or existing.material_qty
            existing.machine_qty = _float(data.get("machine_qty")) or existing.machine_qty
            existing.chapter = chapter_val or existing.chapter
            existing.work_content = _str(data.get("work_content")) or existing.work_content
            existing.applicable_scope = _str(data.get("applicable_scope")) or existing.applicable_scope
            existing.region = _str(data.get("region")) or region or existing.region
            existing.version = version_tag
            existing.profession = profession
            if std_id:
                existing.pricing_standard_id = std_id
            db.add(existing)
            result.updated += 1
            result.items.append(existing)
        else:
            item = QuotaItem(
                quota_code=quota_code,
                name=name,
                unit=_str(data.get("unit")) or "",
                labor_qty=_float(data.get("labor_qty")),
                material_qty=_float(data.get("material_qty")),
                machine_qty=_float(data.get("machine_qty")),
                labor_fee=labor_fee,
                material_fee=material_fee,
                machine_fee=machine_fee,
                base_price=base_price,
                chapter=chapter_val,
                work_content=_str(data.get("work_content")),
                applicable_scope=_str(data.get("applicable_scope")),
                region=_str(data.get("region")) or region,
                version=version_tag,
                profession=profession,
                pricing_standard_id=std_id,
                has_resource_details=0,
                conversion_rules_json="[]",
                unit_constraint_json="{}",
            )
            db.add(item)
            result.imported += 1
            result.items.append(item)

        if rows_processed % 100 == 0:
            db.flush()  # periodic flush for large files

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        result.errors.append(f"数据库提交失败: {exc}")
        logger.error("Quota import commit failed: %s", exc, exc_info=True)

    logger.info(
        "Quota import done: imported=%d updated=%d skipped=%d errors=%d",
        result.imported, result.updated, result.skipped, len(result.errors),
    )
    return result


# ---------------------------------------------------------------------------
# Programmatic bulk-insert helper (for seeding / testing)
# ---------------------------------------------------------------------------

def seed_quota_items(
    items: list[dict],
    db: Session,
    standard_code: str = "GBT50500-2024",
    profession: str = "房建",
    region: str = "全国",
    upsert: bool = True,
) -> QuotaImportResult:
    """Insert/update quota items from a list of dicts.

    Each dict may have keys: quota_code, name, unit, labor_fee, material_fee,
    machine_fee, base_price, chapter, work_content, applicable_scope.
    """
    result = QuotaImportResult()
    std_row = db.query(PricingStandard).filter(PricingStandard.code == standard_code).first()
    std_id: int | None = std_row.id if std_row else None
    version_tag = f"{standard_code}-{profession}"

    for d in items:
        quota_code = str(d.get("quota_code", "")).strip()
        name = str(d.get("name", "")).strip()
        if not quota_code or not name:
            result.skipped += 1
            continue

        existing = db.query(QuotaItem).filter(QuotaItem.quota_code == quota_code).first()
        labor_fee = float(d.get("labor_fee", 0) or 0)
        material_fee = float(d.get("material_fee", 0) or 0)
        machine_fee = float(d.get("machine_fee", 0) or 0)
        base_price = float(d.get("base_price", 0) or 0) or round(labor_fee + material_fee + machine_fee, 2)

        if existing:
            if not upsert:
                result.skipped += 1
                continue
            existing.name = name
            existing.unit = str(d.get("unit", existing.unit))
            existing.labor_fee = labor_fee
            existing.material_fee = material_fee
            existing.machine_fee = machine_fee
            existing.base_price = base_price
            existing.chapter = str(d.get("chapter", existing.chapter))
            existing.work_content = str(d.get("work_content", existing.work_content))
            existing.profession = profession
            existing.region = str(d.get("region", region))
            existing.version = version_tag
            if std_id:
                existing.pricing_standard_id = std_id
            db.add(existing)
            result.updated += 1
            result.items.append(existing)
        else:
            item = QuotaItem(
                quota_code=quota_code,
                name=name,
                unit=str(d.get("unit", "")),
                labor_qty=float(d.get("labor_qty", 0) or 0),
                material_qty=float(d.get("material_qty", 0) or 0),
                machine_qty=float(d.get("machine_qty", 0) or 0),
                labor_fee=labor_fee,
                material_fee=material_fee,
                machine_fee=machine_fee,
                base_price=base_price,
                chapter=str(d.get("chapter", "")),
                work_content=str(d.get("work_content", "")),
                applicable_scope=str(d.get("applicable_scope", "")),
                region=str(d.get("region", region)),
                version=version_tag,
                profession=profession,
                pricing_standard_id=std_id,
                has_resource_details=0,
                conversion_rules_json="[]",
                unit_constraint_json="{}",
            )
            db.add(item)
            result.imported += 1
            result.items.append(item)

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        result.errors.append(f"数据库提交失败: {exc}")

    return result


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _str(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _float(val: object) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0
