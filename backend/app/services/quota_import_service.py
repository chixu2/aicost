"""Service for parsing Excel quota (定额) files and persisting to DB."""

from __future__ import annotations

import io
from dataclasses import dataclass

import openpyxl
from sqlalchemy.orm import Session

from app.models.quota_item import QuotaItem
from app.models.quota_resource_detail import QuotaResourceDetail

_HEADER_MAP: dict[str, str] = {
    "定额号": "quota_code", "定额编号": "quota_code", "quota_code": "quota_code", "code": "quota_code",
    "名称": "name", "定额名称": "name", "name": "name",
    "单位": "unit", "计量单位": "unit", "unit": "unit",
    "人工": "labor_qty", "人工含量": "labor_qty", "labor_qty": "labor_qty", "labor": "labor_qty",
    "材料": "material_qty", "材料含量": "material_qty", "material_qty": "material_qty", "material": "material_qty",
    "机械": "machine_qty", "机械含量": "machine_qty", "machine_qty": "machine_qty", "machine": "machine_qty",
    # Extended fields
    "工作内容": "work_content", "work_content": "work_content",
    "适用范围": "applicable_scope", "applicable_scope": "applicable_scope",
    "章节": "chapter", "chapter": "chapter", "分部": "chapter",
    "版本": "version", "version": "version",
    "基价": "base_price", "base_price": "base_price",
}


@dataclass
class QuotaImportStats:
    imported: int
    skipped: int
    items: list[QuotaItem]


def _normalize(header: str) -> str | None:
    h = header.strip().lower()
    return _HEADER_MAP.get(h)


def parse_and_import_quota(
    file_bytes: bytes,
    db: Session,
) -> QuotaImportStats:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return QuotaImportStats(imported=0, skipped=0, items=[])

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return QuotaImportStats(imported=0, skipped=0, items=[])

    raw_headers = rows[0]
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(raw_headers):
        if cell is None:
            continue
        field = _normalize(str(cell))
        if field:
            col_map[idx] = field

    required = {"quota_code", "name", "unit"}
    if not required.issubset(set(col_map.values())):
        return QuotaImportStats(imported=0, skipped=len(rows) - 1, items=[])

    imported_items: list[QuotaItem] = []
    skipped = 0
    for row in rows[1:]:
        record: dict[str, str | float] = {}
        for idx, field in col_map.items():
            val = row[idx] if idx < len(row) else None
            if val is not None:
                record[field] = val

        quota_code = str(record.get("quota_code", "")).strip()
        name = str(record.get("name", "")).strip()
        unit = str(record.get("unit", "")).strip()
        if not quota_code or not name or not unit:
            skipped += 1
            continue

        def _float(key: str) -> float:
            try:
                return float(record.get(key, 0))
            except (ValueError, TypeError):
                return 0.0

        item = QuotaItem(
            quota_code=quota_code,
            name=name,
            unit=unit,
            labor_qty=_float("labor_qty"),
            material_qty=_float("material_qty"),
            machine_qty=_float("machine_qty"),
            work_content=str(record.get("work_content", "")).strip(),
            applicable_scope=str(record.get("applicable_scope", "")).strip(),
            chapter=str(record.get("chapter", "")).strip(),
            version=str(record.get("version", "")).strip(),
            base_price=_float("base_price"),
        )
        db.add(item)
        imported_items.append(item)

    db.commit()
    for item in imported_items:
        db.refresh(item)

    wb.close()
    return QuotaImportStats(imported=len(imported_items), skipped=skipped, items=imported_items)


# ---------------------------------------------------------------------------
# Resource detail import (定额资源明细)
# ---------------------------------------------------------------------------

_RESOURCE_HEADER_MAP: dict[str, str] = {
    "定额号": "quota_code", "定额编号": "quota_code", "quota_code": "quota_code",
    "类别": "category", "category": "category", "资源类别": "category",
    "资源编码": "resource_code", "resource_code": "resource_code",
    "资源名称": "resource_name", "resource_name": "resource_name", "名称": "resource_name",
    "规格": "spec", "spec": "spec", "规格型号": "spec",
    "单位": "unit", "unit": "unit",
    "消耗量": "quantity", "quantity": "quantity", "含量": "quantity",
    "单价": "unit_price", "unit_price": "unit_price",
    "主材": "is_main_material", "is_main_material": "is_main_material",
}


@dataclass
class ResourceDetailImportStats:
    imported: int
    skipped: int
    quotas_updated: int


def parse_and_import_resource_details(
    file_bytes: bytes,
    db: Session,
) -> ResourceDetailImportStats:
    """Import quota resource details from Excel.

    Expected columns: 定额号, 类别, 资源编码, 资源名称, 规格, 单位, 消耗量, 单价, 主材
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return ResourceDetailImportStats(imported=0, skipped=0, quotas_updated=0)

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return ResourceDetailImportStats(imported=0, skipped=0, quotas_updated=0)

    raw_headers = rows[0]
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(raw_headers):
        if cell is None:
            continue
        h = str(cell).strip().lower()
        field = _RESOURCE_HEADER_MAP.get(h)
        if field:
            col_map[idx] = field

    required = {"quota_code", "category", "resource_name", "unit"}
    if not required.issubset(set(col_map.values())):
        return ResourceDetailImportStats(imported=0, skipped=len(rows) - 1, quotas_updated=0)

    # Pre-load quota lookup
    all_quotas = {q.quota_code: q for q in db.query(QuotaItem).all()}

    imported = 0
    skipped = 0
    updated_quota_ids: set[int] = set()

    for row in rows[1:]:
        record: dict[str, str | float] = {}
        for idx, field in col_map.items():
            val = row[idx] if idx < len(row) else None
            if val is not None:
                record[field] = val

        quota_code = str(record.get("quota_code", "")).strip()
        category = str(record.get("category", "")).strip()
        resource_name = str(record.get("resource_name", "")).strip()
        unit = str(record.get("unit", "")).strip()

        if not quota_code or not category or not resource_name or not unit:
            skipped += 1
            continue

        if category not in ("人工", "材料", "机械"):
            skipped += 1
            continue

        quota = all_quotas.get(quota_code)
        if not quota:
            skipped += 1
            continue

        def _float(key: str) -> float:
            try:
                return float(record.get(key, 0))
            except (ValueError, TypeError):
                return 0.0

        is_main = str(record.get("is_main_material", "")).strip()
        is_main_int = 1 if is_main in ("1", "是", "yes", "true", "Y") else 0

        detail = QuotaResourceDetail(
            quota_item_id=quota.id,
            category=category,
            resource_code=str(record.get("resource_code", "")).strip(),
            resource_name=resource_name,
            spec=str(record.get("spec", "")).strip(),
            unit=unit,
            quantity=_float("quantity"),
            unit_price=_float("unit_price"),
            is_main_material=is_main_int,
        )
        db.add(detail)
        imported += 1
        updated_quota_ids.add(quota.id)

    # Mark quotas as having resource details
    for qid in updated_quota_ids:
        quota = db.query(QuotaItem).filter(QuotaItem.id == qid).first()
        if quota:
            quota.has_resource_details = 1

    db.commit()
    wb.close()
    return ResourceDetailImportStats(
        imported=imported, skipped=skipped, quotas_updated=len(updated_quota_ids),
    )
