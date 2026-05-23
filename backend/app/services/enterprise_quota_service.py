"""Enterprise quota library service: CRUD + state machine + Excel import."""

from __future__ import annotations

import io
import json
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any

from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.enterprise_quota_item import (
    ALL_STATUSES,
    SOURCE_IMPORTED,
    SOURCE_MANUAL,
    STATUS_APPROVED,
    STATUS_ARCHIVED,
    STATUS_DRAFT,
    STATUS_IN_REVIEW,
    STATUS_REJECTED,
    EnterpriseQuotaItem,
)
from app.models.enterprise_quota_candidate import (
    CANDIDATE_PENDING,
    EnterpriseQuotaCandidate,
)


# ─── Custom exceptions ──────────────────────────────────────────────


class EnterpriseQuotaError(Exception):
    """Base for service-layer errors."""


class DuplicateQuotaCodeError(EnterpriseQuotaError):
    pass


class InvalidStateTransition(EnterpriseQuotaError):
    pass


class NotFoundError(EnterpriseQuotaError):
    pass


# ─── Helpers ─────────────────────────────────────────────────────────


def _json_loads(s: str, default: Any) -> Any:
    if not s:
        return default
    try:
        return json.loads(s)
    except (ValueError, TypeError):
        return default


def to_out_dict(item: EnterpriseQuotaItem) -> dict[str, Any]:
    """Serialize ORM object → API dict (matches EnterpriseQuotaOut)."""
    return {
        "id": item.id,
        "quota_code": item.quota_code,
        "name": item.name,
        "unit": item.unit,
        "labor_qty": item.labor_qty,
        "material_qty": item.material_qty,
        "machine_qty": item.machine_qty,
        "labor_fee": item.labor_fee,
        "material_fee": item.material_fee,
        "machine_fee": item.machine_fee,
        "base_price": item.base_price,
        "work_content": item.work_content,
        "applicable_scope": item.applicable_scope,
        "chapter": item.chapter,
        "profession": item.profession,
        "region": item.region,
        "version": item.version,
        "coefficient_default": item.coefficient_default,
        "tags": _json_loads(item.tags_json, []),
        "status": item.status,
        "source_type": item.source_type,
        "source_ref": _json_loads(item.source_ref_json, {}),
        "created_by": item.created_by,
        "created_at": item.created_at,
        "submitted_at": item.submitted_at,
        "reviewed_by": item.reviewed_by,
        "reviewed_at": item.reviewed_at,
        "review_comment": item.review_comment,
        "usage_count": item.usage_count,
    }


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


# ─── CRUD ────────────────────────────────────────────────────────────


def create_item(
    db: Session,
    *,
    data: dict[str, Any],
) -> EnterpriseQuotaItem:
    """Create a new draft enterprise quota item."""
    quota_code = (data.get("quota_code") or "").strip()
    if not quota_code:
        raise EnterpriseQuotaError("quota_code is required")

    existing = (
        db.query(EnterpriseQuotaItem)
        .filter(EnterpriseQuotaItem.quota_code == quota_code)
        .first()
    )
    if existing:
        raise DuplicateQuotaCodeError(f"企业定额编码 '{quota_code}' 已存在")

    tags = data.get("tags") or []
    source_ref = data.get("source_ref") or {}

    item = EnterpriseQuotaItem(
        quota_code=quota_code,
        name=str(data.get("name", "")).strip(),
        unit=str(data.get("unit", "")).strip(),
        labor_qty=float(data.get("labor_qty", 0) or 0),
        material_qty=float(data.get("material_qty", 0) or 0),
        machine_qty=float(data.get("machine_qty", 0) or 0),
        labor_fee=float(data.get("labor_fee", 0) or 0),
        material_fee=float(data.get("material_fee", 0) or 0),
        machine_fee=float(data.get("machine_fee", 0) or 0),
        base_price=float(data.get("base_price", 0) or 0),
        work_content=str(data.get("work_content", "") or ""),
        applicable_scope=str(data.get("applicable_scope", "") or ""),
        chapter=str(data.get("chapter", "") or ""),
        profession=str(data.get("profession", "房建") or "房建"),
        region=str(data.get("region", "") or ""),
        version=str(data.get("version", "v1.0") or "v1.0"),
        coefficient_default=float(data.get("coefficient_default", 1.0) or 1.0),
        tags_json=json.dumps(tags, ensure_ascii=False),
        status=STATUS_DRAFT,
        source_type=str(data.get("source_type") or SOURCE_MANUAL),
        source_ref_json=json.dumps(source_ref, ensure_ascii=False),
        created_by=str(data.get("created_by") or ""),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def update_item(
    db: Session,
    *,
    item_id: int,
    data: dict[str, Any],
) -> EnterpriseQuotaItem:
    """Update mutable fields. Approved/archived items cannot be edited."""
    item = db.query(EnterpriseQuotaItem).filter(EnterpriseQuotaItem.id == item_id).first()
    if not item:
        raise NotFoundError(f"企业定额 {item_id} 不存在")

    if item.status in (STATUS_APPROVED, STATUS_ARCHIVED):
        raise InvalidStateTransition(
            f"已发布或已归档的条目不能直接编辑，请先 restore 到草稿（当前状态: {item.status}）",
        )

    mutable_fields = {
        "name", "unit",
        "labor_qty", "material_qty", "machine_qty",
        "labor_fee", "material_fee", "machine_fee", "base_price",
        "work_content", "applicable_scope", "chapter",
        "profession", "region", "version", "coefficient_default",
    }
    for field in mutable_fields:
        if field in data and data[field] is not None:
            setattr(item, field, data[field])

    if "tags" in data and data["tags"] is not None:
        item.tags_json = json.dumps(data["tags"], ensure_ascii=False)

    db.commit()
    db.refresh(item)
    return item


def delete_item(db: Session, *, item_id: int) -> None:
    """Only draft items may be deleted; others should be archived."""
    item = db.query(EnterpriseQuotaItem).filter(EnterpriseQuotaItem.id == item_id).first()
    if not item:
        raise NotFoundError(f"企业定额 {item_id} 不存在")
    if item.status != STATUS_DRAFT:
        raise InvalidStateTransition(
            f"仅草稿状态可删除（当前状态: {item.status}），请改用归档",
        )
    db.delete(item)
    db.commit()


def list_items(
    db: Session,
    *,
    status: str | None = None,
    source_type: str | None = None,
    keyword: str | None = None,
    profession: str | None = None,
    chapter: str | None = None,
    skip: int = 0,
    limit: int = 50,
) -> tuple[int, list[EnterpriseQuotaItem]]:
    q = db.query(EnterpriseQuotaItem)
    if status:
        q = q.filter(EnterpriseQuotaItem.status == status)
    if source_type:
        q = q.filter(EnterpriseQuotaItem.source_type == source_type)
    if profession:
        q = q.filter(EnterpriseQuotaItem.profession == profession)
    if chapter:
        q = q.filter(EnterpriseQuotaItem.chapter == chapter)
    if keyword:
        kw = f"%{keyword}%"
        q = q.filter(
            (EnterpriseQuotaItem.name.like(kw))
            | (EnterpriseQuotaItem.quota_code.like(kw)),
        )
    total = q.count()
    rows = (
        q.order_by(EnterpriseQuotaItem.id.desc())
        .offset(skip)
        .limit(limit)
        .all()
    )
    return total, rows


def get_item(db: Session, *, item_id: int) -> EnterpriseQuotaItem:
    item = db.query(EnterpriseQuotaItem).filter(EnterpriseQuotaItem.id == item_id).first()
    if not item:
        raise NotFoundError(f"企业定额 {item_id} 不存在")
    return item


# ─── State machine ───────────────────────────────────────────────────


_TRANSITIONS: dict[str, set[str]] = {
    STATUS_DRAFT: {STATUS_IN_REVIEW},
    STATUS_IN_REVIEW: {STATUS_APPROVED, STATUS_REJECTED},
    STATUS_APPROVED: {STATUS_ARCHIVED, STATUS_DRAFT},
    STATUS_REJECTED: {STATUS_DRAFT, STATUS_ARCHIVED},
    STATUS_ARCHIVED: {STATUS_DRAFT},
}


def _transition(item: EnterpriseQuotaItem, target: str) -> None:
    if target not in ALL_STATUSES:
        raise InvalidStateTransition(f"未知状态: {target}")
    allowed = _TRANSITIONS.get(item.status, set())
    if target not in allowed:
        raise InvalidStateTransition(
            f"不允许从 {item.status} 转换到 {target}",
        )
    item.status = target


def submit_for_review(db: Session, *, item_id: int, actor: str = "") -> EnterpriseQuotaItem:
    item = get_item(db, item_id=item_id)
    _transition(item, STATUS_IN_REVIEW)
    item.submitted_at = _utcnow()
    if actor and not item.created_by:
        item.created_by = actor
    db.commit()
    db.refresh(item)
    return item


def approve(db: Session, *, item_id: int, actor: str = "", comment: str = "") -> EnterpriseQuotaItem:
    item = get_item(db, item_id=item_id)
    _transition(item, STATUS_APPROVED)
    item.reviewed_by = actor
    item.reviewed_at = _utcnow()
    item.review_comment = comment
    db.commit()
    db.refresh(item)
    return item


def reject(db: Session, *, item_id: int, actor: str = "", comment: str = "") -> EnterpriseQuotaItem:
    item = get_item(db, item_id=item_id)
    _transition(item, STATUS_REJECTED)
    item.reviewed_by = actor
    item.reviewed_at = _utcnow()
    item.review_comment = comment
    db.commit()
    db.refresh(item)
    return item


def archive(db: Session, *, item_id: int, actor: str = "") -> EnterpriseQuotaItem:
    item = get_item(db, item_id=item_id)
    _transition(item, STATUS_ARCHIVED)
    item.reviewed_by = actor or item.reviewed_by
    item.reviewed_at = _utcnow()
    db.commit()
    db.refresh(item)
    return item


def restore_to_draft(db: Session, *, item_id: int, actor: str = "") -> EnterpriseQuotaItem:
    item = get_item(db, item_id=item_id)
    _transition(item, STATUS_DRAFT)
    item.reviewed_by = actor or item.reviewed_by
    db.commit()
    db.refresh(item)
    return item


# ─── Stats ───────────────────────────────────────────────────────────


def stats(db: Session) -> dict[str, Any]:
    by_status_raw = (
        db.query(EnterpriseQuotaItem.status, func.count())
        .group_by(EnterpriseQuotaItem.status)
        .all()
    )
    by_source_raw = (
        db.query(EnterpriseQuotaItem.source_type, func.count())
        .group_by(EnterpriseQuotaItem.source_type)
        .all()
    )
    by_status = {s: int(n) for s, n in by_status_raw}
    by_source = {s: int(n) for s, n in by_source_raw}
    total = sum(by_status.values())

    pending_candidates = (
        db.query(EnterpriseQuotaCandidate)
        .filter(EnterpriseQuotaCandidate.status == CANDIDATE_PENDING)
        .count()
    )

    cutoff = _utcnow() - timedelta(days=30)
    recent_created = (
        db.query(EnterpriseQuotaItem)
        .filter(EnterpriseQuotaItem.created_at >= cutoff)
        .count()
    )

    return {
        "total": total,
        "by_status": by_status,
        "by_source": by_source,
        "pending_review": by_status.get(STATUS_IN_REVIEW, 0),
        "pending_candidates": pending_candidates,
        "recent_created": recent_created,
    }


# ─── Excel import ────────────────────────────────────────────────────


_HEADER_MAP: dict[str, str] = {
    "定额号": "quota_code", "定额编号": "quota_code", "编码": "quota_code",
    "名称": "name", "定额名称": "name",
    "单位": "unit", "计量单位": "unit",
    "人工含量": "labor_qty", "人工": "labor_qty",
    "材料含量": "material_qty", "材料": "material_qty",
    "机械含量": "machine_qty", "机械": "machine_qty",
    "人工费": "labor_fee", "材料费": "material_fee", "机械费": "machine_fee",
    "基价": "base_price",
    "工作内容": "work_content",
    "适用范围": "applicable_scope",
    "章节": "chapter", "分部": "chapter",
    "专业": "profession",
    "地区": "region",
    "版本": "version",
    "默认系数": "coefficient_default", "系数": "coefficient_default",
}


@dataclass
class ImportResult:
    imported: int = 0
    skipped: int = 0
    errors: list[str] = None  # type: ignore[assignment]


def _normalize_header(h: str) -> str | None:
    return _HEADER_MAP.get(str(h).strip())


def import_from_excel(
    file_bytes: bytes,
    db: Session,
    *,
    created_by: str = "",
) -> ImportResult:
    """Parse Excel file, create draft items. Skips duplicates by quota_code."""
    import openpyxl

    result = ImportResult(imported=0, skipped=0, errors=[])

    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        result.errors.append(f"无法打开 Excel 文件: {e}")
        return result

    ws = wb.active
    if ws is None:
        result.errors.append("Excel 工作表为空")
        return result

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        result.errors.append("Excel 至少需要表头 + 一行数据")
        return result

    headers = rows[0]
    col_map: dict[int, str] = {}
    for idx, cell in enumerate(headers):
        if cell is None:
            continue
        field = _normalize_header(str(cell))
        if field:
            col_map[idx] = field

    required = {"quota_code", "name", "unit"}
    if not required.issubset(set(col_map.values())):
        result.errors.append("缺少必需列: 定额号 / 名称 / 单位")
        return result

    for row_idx, row in enumerate(rows[1:], start=2):
        record: dict[str, Any] = {}
        for col_idx, field_name in col_map.items():
            if col_idx < len(row):
                val = row[col_idx]
                if val is not None:
                    record[field_name] = val

        quota_code = str(record.get("quota_code", "")).strip()
        name = str(record.get("name", "")).strip()
        unit = str(record.get("unit", "")).strip()
        if not quota_code or not name or not unit:
            result.skipped += 1
            continue

        # Skip duplicates
        existing = (
            db.query(EnterpriseQuotaItem)
            .filter(EnterpriseQuotaItem.quota_code == quota_code)
            .first()
        )
        if existing:
            result.skipped += 1
            continue

        def _f(key: str, default: float = 0.0) -> float:
            try:
                return float(record.get(key, default) or default)
            except (ValueError, TypeError):
                return default

        try:
            item = EnterpriseQuotaItem(
                quota_code=quota_code,
                name=name,
                unit=unit,
                labor_qty=_f("labor_qty"),
                material_qty=_f("material_qty"),
                machine_qty=_f("machine_qty"),
                labor_fee=_f("labor_fee"),
                material_fee=_f("material_fee"),
                machine_fee=_f("machine_fee"),
                base_price=_f("base_price"),
                work_content=str(record.get("work_content", "") or ""),
                applicable_scope=str(record.get("applicable_scope", "") or ""),
                chapter=str(record.get("chapter", "") or ""),
                profession=str(record.get("profession", "房建") or "房建"),
                region=str(record.get("region", "") or ""),
                version=str(record.get("version", "v1.0") or "v1.0"),
                coefficient_default=_f("coefficient_default", 1.0) or 1.0,
                tags_json="[]",
                status=STATUS_DRAFT,
                source_type=SOURCE_IMPORTED,
                source_ref_json=json.dumps(
                    {"imported_at": _utcnow().isoformat(), "row": row_idx},
                    ensure_ascii=False,
                ),
                created_by=created_by,
            )
            db.add(item)
            result.imported += 1
        except Exception as e:  # noqa: BLE001
            result.skipped += 1
            result.errors.append(f"第 {row_idx} 行: {e}")

    db.commit()
    return result


# ─── Excel template generation ───────────────────────────────────────


def build_template_xlsx() -> bytes:
    """Return an .xlsx template (header + sample rows)."""
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:  # safety
        return b""
    ws.title = "企业定额导入模板"

    headers = [
        "定额号", "名称", "单位",
        "人工含量", "材料含量", "机械含量",
        "人工费", "材料费", "机械费", "基价",
        "工作内容", "适用范围", "章节", "专业", "地区",
        "版本", "默认系数",
    ]
    sample = [
        "ENT-A01001", "C30 现浇混凝土柱（企业经验）", "m³",
        0.85, 1.05, 0.12,
        128.50, 415.20, 18.40, 562.10,
        "包含模板支拆、混凝土浇筑、振捣、养护",
        "适用于一般工业与民用建筑现浇混凝土柱",
        "混凝土工程", "房建", "全国",
        "v2026.1", 1.0,
    ]
    sample2 = [
        "ENT-A02003", "240 厚加气混凝土砌块墙（企业经验）", "m³",
        0.55, 1.02, 0.05,
        82.50, 178.40, 6.20, 267.10,
        "包含砌筑、灰缝处理、构造柱拉结",
        "外墙、内隔墙均可使用",
        "砌筑工程", "房建", "全国",
        "v2026.1", 1.0,
    ]

    header_fill = PatternFill("solid", fgColor="1d6fe8")
    header_font = Font(name="微软雅黑", size=11, color="FFFFFF", bold=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = 15

    for col_idx, val in enumerate(sample, 1):
        ws.cell(row=2, column=col_idx, value=val)
    for col_idx, val in enumerate(sample2, 1):
        ws.cell(row=3, column=col_idx, value=val)

    # Notes
    ws.cell(row=5, column=1, value="说明:")
    ws.cell(row=6, column=1, value="• 必填: 定额号 / 名称 / 单位")
    ws.cell(row=7, column=1, value="• 含量字段(人工/材料/机械含量)与费用字段(人工/材料/机械费)二选一即可")
    ws.cell(row=8, column=1, value="• 默认系数若不填则按 1.0 处理")
    ws.cell(row=9, column=1, value="• 导入后条目状态为草稿(draft)，需提交审批后方可使用")

    bio = io.BytesIO()
    wb.save(bio)
    wb.close()
    return bio.getvalue()
