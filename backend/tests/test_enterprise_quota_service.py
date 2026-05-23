"""Tests for enterprise_quota_service: CRUD + state machine + import."""

import io

import openpyxl
import pytest

from app.models.enterprise_quota_item import (
    SOURCE_IMPORTED,
    SOURCE_MANUAL,
    STATUS_APPROVED,
    STATUS_DRAFT,
    STATUS_IN_REVIEW,
    STATUS_REJECTED,
    EnterpriseQuotaItem,
)
from app.services import enterprise_quota_service as svc


def _payload(**overrides):
    base = {
        "quota_code": "ENT-A001",
        "name": "测试项",
        "unit": "m³",
        "labor_qty": 0.5,
        "material_qty": 1.2,
        "machine_qty": 0.1,
        "base_price": 300.0,
        "coefficient_default": 1.0,
        "tags": ["test"],
        "created_by": "alice",
    }
    base.update(overrides)
    return base


# ─── CRUD ────────────────────────────────────────────────────────────


def test_create_and_get(db):
    item = svc.create_item(db, data=_payload())
    assert item.id and item.status == STATUS_DRAFT
    assert item.source_type == SOURCE_MANUAL
    fetched = svc.get_item(db, item_id=item.id)
    assert fetched.quota_code == "ENT-A001"


def test_create_duplicate_code_raises(db):
    svc.create_item(db, data=_payload())
    with pytest.raises(svc.DuplicateQuotaCodeError):
        svc.create_item(db, data=_payload())


def test_update_draft(db):
    item = svc.create_item(db, data=_payload())
    updated = svc.update_item(
        db, item_id=item.id,
        data={"name": "改名", "base_price": 400.0, "tags": ["a", "b"]},
    )
    assert updated.name == "改名"
    assert updated.base_price == 400.0
    assert "a" in updated.tags_json


def test_cannot_edit_approved(db):
    item = svc.create_item(db, data=_payload())
    svc.submit_for_review(db, item_id=item.id)
    svc.approve(db, item_id=item.id, actor="bob")
    with pytest.raises(svc.InvalidStateTransition):
        svc.update_item(db, item_id=item.id, data={"name": "X"})


def test_delete_only_when_draft(db):
    item = svc.create_item(db, data=_payload())
    svc.submit_for_review(db, item_id=item.id)
    with pytest.raises(svc.InvalidStateTransition):
        svc.delete_item(db, item_id=item.id)
    # Reject → back to draft → deletable via restore
    svc.reject(db, item_id=item.id)
    svc.restore_to_draft(db, item_id=item.id)
    svc.delete_item(db, item_id=item.id)
    assert db.query(EnterpriseQuotaItem).count() == 0


def test_list_with_filters(db):
    svc.create_item(db, data=_payload(quota_code="ENT-A001", chapter="混凝土"))
    svc.create_item(db, data=_payload(quota_code="ENT-A002", chapter="砌筑", name="砌墙"))
    total, rows = svc.list_items(db, chapter="混凝土")
    assert total == 1
    total2, rows2 = svc.list_items(db, keyword="砌")
    assert total2 == 1 and rows2[0].quota_code == "ENT-A002"


# ─── State machine ───────────────────────────────────────────────────


def test_full_workflow(db):
    item = svc.create_item(db, data=_payload())
    assert item.status == STATUS_DRAFT

    item = svc.submit_for_review(db, item_id=item.id, actor="alice")
    assert item.status == STATUS_IN_REVIEW
    assert item.submitted_at is not None

    item = svc.approve(db, item_id=item.id, actor="bob", comment="LGTM")
    assert item.status == STATUS_APPROVED
    assert item.reviewed_by == "bob"
    assert item.review_comment == "LGTM"


def test_reject_then_restore(db):
    item = svc.create_item(db, data=_payload())
    svc.submit_for_review(db, item_id=item.id)
    svc.reject(db, item_id=item.id, actor="bob", comment="价格偏高")
    item = svc.get_item(db, item_id=item.id)
    assert item.status == STATUS_REJECTED

    svc.restore_to_draft(db, item_id=item.id)
    item = svc.get_item(db, item_id=item.id)
    assert item.status == STATUS_DRAFT


def test_invalid_transition_raises(db):
    item = svc.create_item(db, data=_payload())
    # Cannot directly approve a draft
    with pytest.raises(svc.InvalidStateTransition):
        svc.approve(db, item_id=item.id)


# ─── Stats ───────────────────────────────────────────────────────────


def test_stats(db):
    a = svc.create_item(db, data=_payload(quota_code="A"))
    svc.create_item(db, data=_payload(quota_code="B"))
    svc.submit_for_review(db, item_id=a.id)

    s = svc.stats(db)
    assert s["total"] == 2
    assert s["by_status"][STATUS_DRAFT] == 1
    assert s["by_status"][STATUS_IN_REVIEW] == 1
    assert s["pending_review"] == 1


# ─── Excel import ────────────────────────────────────────────────────


def _build_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [
        "定额号", "名称", "单位",
        "人工含量", "材料含量", "机械含量",
        "基价", "工作内容", "章节", "默认系数",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_import_from_excel(db):
    data = _build_excel([
        ["EI-001", "C30 混凝土", "m³", 0.6, 1.0, 0.05, 350.0, "包含浇筑", "混凝土", 1.0],
        ["EI-002", "钢筋", "t", 12.0, 0, 0, 5500.0, "", "钢筋", 1.0],
    ])
    result = svc.import_from_excel(data, db, created_by="importer")
    assert result.imported == 2
    assert result.skipped == 0

    items = db.query(EnterpriseQuotaItem).all()
    assert len(items) == 2
    assert all(it.source_type == SOURCE_IMPORTED for it in items)
    assert all(it.status == STATUS_DRAFT for it in items)


def test_import_skips_duplicates(db):
    svc.create_item(db, data=_payload(quota_code="EI-001"))
    data = _build_excel([
        ["EI-001", "重复", "m³", 0, 0, 0, 0, "", "", 1.0],  # dup
        ["EI-002", "新条目", "m³", 0, 0, 0, 0, "", "", 1.0],
    ])
    result = svc.import_from_excel(data, db)
    assert result.imported == 1
    assert result.skipped == 1


def test_template_download_returns_xlsx(db):  # noqa: ARG001
    data = svc.build_template_xlsx()
    assert data.startswith(b"PK")  # zip/xlsx signature
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert "定额号" in headers and "名称" in headers and "单位" in headers
